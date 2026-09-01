# -*- coding: utf-8 -*-
"""
카카오 비즈니스 채널 콘솔에서 "엑셀로 내보내기" 한 대화 파일(.xlsx)을 파싱한다.

파일 형식: 시트 이름 = "{채널명}_{고객닉네임}", 컬럼 = DATE, USER, MESSAGE.
USER가 채널명(에이전트)이면 상담원 메시지, 그 외엔 고객 메시지로 취급한다.

읽어온 셀 값에 엑셀 XML 이스케이프가 덜 풀린 "_x000D_"(캐리지리턴) 잔재가 섞여 있는
경우가 있어 제거하고, "지금은 ~ 채팅 가능한 시간이 아닙니다" 같은 자동 응답 문구는
실제 상담 내용이 아니라서 걸러낸다.
"""
import io
import re
import openpyxl

AGENT_NAME = "더비랩"
_AUTO_REPLY_PATTERNS = [
    "채팅 운영시간 안내",
    "채팅 가능한 시간이 아닙니다",
]


def _clean(text):
    if not text:
        return ""
    return str(text).replace("_x000D_", "").strip()


def _is_auto_reply(text):
    return any(p in text for p in _AUTO_REPLY_PATTERNS)


def parse_kakao_export(file_bytes):
    """반환: 사람이 읽기 좋은 "역할: 메시지" 형태로 합친 대화 전문(문자열).
    파싱 실패/빈 파일이면 빈 문자열."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    target_sheet = None
    for name in wb.sheetnames:
        ws = wb[name]
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        if header[:3] == ["DATE", "USER", "MESSAGE"]:
            target_sheet = ws
            break
    if target_sheet is None:
        return ""

    lines = []
    for row in target_sheet.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 3:
            continue
        date_val, user, message = row[0], _clean(row[1]), _clean(row[2])
        if not message or _is_auto_reply(message):
            continue
        role = "상담원" if user == AGENT_NAME else "고객"
        time_str = ""
        if date_val is not None:
            try:
                time_str = date_val.strftime("%m/%d %H:%M")
            except AttributeError:
                time_str = str(date_val)
        prefix = f"[{time_str}] " if time_str else ""
        lines.append(f"{prefix}{role}: {message}")

    return "\n".join(lines)
